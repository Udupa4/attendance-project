import json
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Load users and attendance data
with open('users.json', 'r') as users_file:
    users_data = json.load(users_file)

with open('attendance.json', 'r') as attendance_file:
    attendance_data = json.load(attendance_file)

# Extract unique dates from attendance data
unique_dates = sorted(set(value.split()[0] for value in attendance_data.values()))

# print(unique_dates)

# Load or create Excel workbook
try:
    workbook = openpyxl.load_workbook('attendance.xlsx')
except FileNotFoundError:
    workbook = Workbook()

# Select active worksheet
worksheet = workbook.active

# Add a new column with today's date
for date in unique_dates:
    new_column_letter = get_column_letter(worksheet.max_column + 1)
    new_column_name = date
    worksheet[new_column_letter + '1'] = new_column_name

# Iterate through all users
for user_id, status in users_data.items():
    # Initialize a row with user_id
    row = [user_id]
    # Fill attendance data for each date
    for date in unique_dates:
        attendance_datetime = attendance_data.get(user_id)
        if attendance_datetime is not None:
            # Extract date part and compare
            if attendance_datetime.split()[0] == date:
                # Extract time and append to row
                row.append(attendance_datetime.split()[1])
                continue
        # If user not present on this date, mark as absent
        row.append("absent")
    # Append the row to the worksheet
    worksheet.append(row)

# Save the workbook
workbook.save('attendance.xlsx')
